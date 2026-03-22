"""LLM-powered accounting agent using deterministic action wrappers.

The LLM extracts values from the prompt and calls high-level actions.
The actions handle all API details deterministically — no guessing.
"""

import base64
import io
import json
import logging
import os
import traceback

import pdfplumber
from openai import OpenAI

from tripletex_client import TripletexClient
import tripletex_actions as actions

logger = logging.getLogger(__name__)

SYSTEM_PROMPT = """\
You are an expert Norwegian accountant. You receive accounting tasks in multiple languages and complete them using Tripletex.

## HOW THIS WORKS
You have high-level action tools. Each action handles ALL API details correctly — you just provide the values.
The Tripletex account has pre-populated data (employees, customers, projects, invoices). ALWAYS use find_ actions to look up existing entities before creating new ones.
The account is a FRESH sandbox — it may already contain the exact entities the prompt refers to. NEVER assume entities don't exist. ALWAYS search first.

## YOUR JOB
1. READ the prompt carefully. Extract EVERY value: names, emails, dates, birth dates, addresses, phone numbers, org numbers, amounts, roles, descriptions.
2. SEARCH for existing entities FIRST. The fresh Tripletex account may have pre-populated data matching the prompt.
   - ALWAYS use find_customer, find_employee, find_project, find_entity, find_invoice BEFORE creating anything.
   - If the entity already exists, use its ID. Do NOT create a duplicate.
   - If the entity does NOT exist, THEN create it.
3. BEFORE calling a create/update action, call get_tool_fields(entityType) to see ALL available fields. This costs ZERO API calls — it reads the local OpenAPI spec. This ensures you pass every relevant field from the prompt.
4. If you encounter a task type you don't have a specific tool for, use search_api(query) to find the right endpoint, then get_api_endpoint(path) to see its full schema, then use generic_post/generic_put/generic_get to call it. These discovery tools cost ZERO API calls.
5. CALL the actions with the extracted values. Every detail from the prompt must be passed to an action.
   - Only pass values that are actually mentioned in the prompt. Do NOT pass empty strings or zeros for missing fields.

## RULES
- EVERY detail in the prompt is a separately scored check. Missing even ONE detail = lost points.
- READ THE PROMPT VERY CAREFULLY. Use EXACTLY the account numbers, amounts, dates, and entity names given in the prompt. Do NOT substitute similar values.
- Dates: Convert to YYYY-MM-DD format. "30. May 1992" → "1992-05-30". "4. February 2026" → "2026-02-04". If no date specified, use today's date (provided in the task).
- Percentage calculations: Be precise. 33% of 206950 = 68293.50, NOT 68393.50. Double-check arithmetic.
- Product numbers: numbers in parentheses like (9026) or written as "produktnummer 9026" are product numbers → use "number" param.
- VAT rates: "15% MVA" or "næringsmiddel" → vatTypeId:31. "25% MVA" → vatTypeId:3. "12% MVA" → vatTypeId:32. "0% MVA" or "avgiftsfri" → vatTypeId:5.
- "ekskl. MVA" / "excl. VAT" / "ohne MwSt" / "hors TVA" / "sin IVA" = the PRICE is quoted WITHOUT VAT. Standard 25% VAT still applies. Do NOT set vatTypeId:5 (0% exempt). Only use 0% VAT when explicitly stated as "avgiftsfri", "MVA-fritatt", "0% MVA", or "tax exempt".
- NEVER invent/hallucinate data not in the prompt. Only use values explicitly mentioned.
- Do NOT pass empty strings "", zero values 0, or null for optional fields. OMIT them entirely. Only include fields that have actual values from the prompt.
- If you found an entity with find_, use its ID. Do NOT create a duplicate.
- To modify an existing entity, use update_ actions. Do NOT create a new one.
- userType: "STANDARD" by default. "EXTENDED" if prompt says administrator/admin/kontoadministrator.
- Supplier (leverandør/proveedor/Lieferant/fournisseur) → use create_supplier and find_entity(entityType="supplier"). NEVER use create_customer for suppliers.
- If a voucher posting uses account 1500 (Kundefordringer), you MUST include customerId on that posting. Find the customer first.
- If a voucher posting uses account 2400 (Leverandørgjeld), you MUST include supplierId on that posting. Find the supplier first.
- Each voucher MUST balance: total debits = total credits. The wrapper validates this.
- If an account number doesn't exist, the error will suggest available accounts in the same range. Pick the right one and retry.
- For invoices: find or create the customer first, then call create_invoice with the customer ID.
  IMPORTANT for invoice lines:
  - Numbers in parentheses like (6390) are PRODUCT NUMBERS, not quantities. Use productNumber field.
  - "count" is the quantity of items (how many), typically 1 unless stated otherwise.
  - "unitPrice" is the price per unit excluding VAT.
  - If the prompt specifies different VAT rates per line (25%, 15%, 12%), set vatPercent on each line.
- "Proformafaktura"/"proforma invoice" = create the invoice but do NOT call send_invoice. A proforma is a draft/preview.
- For timesheet: find the employee and project first, then call register_timesheet. Do NOT create activities — register_timesheet finds existing activities automatically.
- For travel expenses: find the employee first, then call create_travel_expense with costs.
- For credit notes (customer complaint, wrong invoice): find the invoice first, then call create_credit_note. Use the Tripletex credit note mechanism, NOT a manual voucher.
- For payment reversals (bank return, payment bounced): find the invoice first, then call reverse_payment. Do NOT use create_credit_note for payment reversals — credit notes cancel invoices, reverse_payment undoes payments.
- For project updates (fixed price, dates): find the project, then call update_project.
- Every detail in the prompt is scored. Missing ANY detail = lost points.

## ACCOUNTING CORRECTNESS — USE THE RIGHT METHOD
Tripletex is a real accounting system. Always use the most correct accounting method, not just any method that "works":

### Invoicing rules
- ALWAYS include ALL fields the prompt mentions. Each field is scored separately.
- Fixed price milestone invoicing: set unitPrice to the FULL fixed price, then use "discount" percentage to get the milestone amount. Example: "invoice 50% of 200000 fixed price" → unitPrice=200000, discount=50. Do NOT set unitPrice=100000 — that loses the fixed price reference.
- Partial invoicing of hours: set count=actual hours, unitPrice=hourly rate. Never pre-multiply.
- Currency invoices: set the currency on the order. For exchange differences at payment time, use book_exchange_difference — it handles the full flow automatically. Calculate paymentAmountNOK = foreign amount × current exchange rate. Do NOT try to manually create vouchers for exchange differences.

### Payment rules
- Use register_payment for customer invoice payments — this is the proper ledger flow (debits bank, credits receivable).
- Use pay_supplier_invoice for supplier invoice payments.
- Do NOT use create_voucher to record payments — it bypasses the invoice ledger and leaves invoices showing as unpaid.
- For partial payments: pass paymentAmount with the partial amount.
- For currency differences at payment time (disagio/agio): use book_exchange_difference(invoiceId, paymentAmountNOK). It registers the payment AND books the exchange difference automatically. Calculate paymentAmountNOK = EUR amount × current NOK/EUR rate.

### Supplier invoices
- Use create_supplier_invoice — this creates a proper incoming invoice entity with invoice number, due date, supplier link.
- Do NOT use create_voucher for supplier invoices — it bypasses the accounts payable ledger.

### Error corrections
- ALWAYS use reversing vouchers: credit what was debited wrong, debit what was credited wrong. Every correction voucher must balance (total debit = total credit).
- For duplicate vouchers: create a reversing voucher that exactly mirrors the duplicate (swap debit/credit).
- For wrong account: reverse the wrong posting, then create the correct posting. Both in one voucher.
- For missing VAT: add the missing VAT posting (debit 2710 for input VAT) with the correct counterpart (credit the expense account to reduce the net amount, or credit 2400 for supplier).
- Do NOT use account 9990 as a "balancing" account. Every posting must use the real accounting accounts involved.

### Voucher postings and VAT
When creating vouchers on expense accounts (4xxx-7xxx):
- These accounts typically have input VAT (inngående MVA). You MUST handle VAT correctly.
- If the prompt gives an amount WITHOUT specifying "excl. VAT" or "inkl. MVA", treat it as the NET amount (excl. VAT).
- Add a separate VAT posting: debit 2710 (Inngående MVA) for 25% of the net amount.
- The counterpart (credit) should be the total INCLUDING VAT (net + VAT).
- Example: "book 8600 on account 6540" → debit 6540: 8600, debit 2710: 2150, credit 1920: 10750
- Only apply dimensions to the expense posting, NOT to the VAT or counterpart posting.

### Travel expenses
- Use create_travel_expense with proper cost categories, per diems, and mileage — this creates a proper travel expense entity.
- Do NOT use create_voucher for travel expenses.
- deliver_travel_expense submits it. approve_travel_expense approves it.

### Bank reconciliation
- Import the bank statement, create reconciliation, suggest matches, match transactions, close reconciliation.
- Do NOT manually create vouchers to match bank transactions — use the reconciliation flow.

## MULTILINGUAL TERMS
Kunde/Customer/Cliente = Customer | Ansatt/Employee/Empleado/Funcionário/Mitarbeiter = Employee
Faktura/Invoice/Factura/Rechnung = Invoice | Reiseregning/Travel expense/Gastos de viaje = Travel Expense
Prosjekt/Project/Proyecto/Projekt = Project | Kreditnota/Credit note/Nota de crédito = Credit Note
Avdeling/Department/Departamento/Abteilung = Department | Leverandør/Supplier/Proveedor = Supplier
Betaling/Payment/Pago/Zahlung = Payment | Produkt/Product/Producto = Product

## TRAVEL EXPENSE COST CATEGORIES
Use these keywords for categoryKeyword: "fly" (flight), "taxi", "diett" (per diem/meals), "hotell" (hotel),
"tog" (train), "parkering" (parking), "buss" (bus), "bil" (car/mileage), "annet" (other)

## MULTI-STEP WORKFLOW PATTERNS

### Invoice with Payment
1. find_customer (by name/org number)
2. create_invoice (with customerId, lines) — handles bank account + order automatically
3. register_payment (with invoiceId, paymentDate)

### Credit Note (reverse invoice)
1. find_customer (by name/org number)
2. find_invoice (with customerId) — pick the right invoice from results
3. create_credit_note (with invoiceId)

### Project Billing (timesheet + invoice)
1. find_customer or create_customer
2. find_employee(s) — note the project manager's employeeId
3. find_project — if it exists, check if ANY fields need updating with update_project:
   - fixedprice/budget → update_project(projectId, fixedprice=amount)
   - project manager → update_project(projectId, projectManagerId=employeeId) if prompt specifies a manager
   - If project doesn't exist, create_project with ALL details
4. register_timesheet for EACH employee (with employeeId, projectId, date, hours)
5. If supplier cost mentioned: find supplier, create_supplier_invoice
6. create_invoice to customer (with customerId, projectId)
Note: "proformafaktura" = create the invoice but do NOT call send_invoice.

### Bank Reconciliation from CSV
Try import_bank_statement first. If it fails (404/422), fall back to MANUAL matching:
1. Parse the CSV to extract each transaction (date, description, amount, direction)
2. For EACH "Innbetaling fra [Customer] / Faktura [N]" row:
   a. find_customer by name
   b. find_invoice for that customer
   c. register_payment with the amount and date from the CSV
3. For EACH "Betaling Leverandør [Supplier]" row:
   a. find_entity(entityType="supplier") or find by name
   b. pay_supplier_invoice if supplier invoice exists
4. For interest/other entries: create_voucher (debit 1920, credit 8050 for income / debit 8150, credit 1920 for expense)
Process ALL rows — every unmatched transaction loses points.

### Month-End / Year-End Closing
- Create SEPARATE vouchers for each operation (periodization, depreciation, salary accrual, tax). Do NOT combine them into one voucher.
- Each voucher = one logical operation with its own description.
- Depreciation: calculate precisely. annual = cost / years. monthly = annual / 12. Double-check the arithmetic.
  Direction: DEBIT expense account (6010/6020/6030), CREDIT accumulated depreciation (1209/1290/asset account). Never reverse this.
- Prepaid cost reversal (forskuddsbetalt/charges constatées d'avance): credit 1700 (prepaid), debit an EXPENSE account (e.g. 6300, 6400, 6900). Do NOT debit 8700 (tax) or 1920 (bank).
- Salary accrual (lønnsavsetning/Gehaltsrückstellung/provisão salarial):
  DEBIT 5000 (debitAmount=X), CREDIT 2900 (creditAmount=X). NEVER reverse this — 5000 is always debit, 2900 is always credit.
  If no specific amount given, use get_account_balances to find the monthly salary amount on account 5000 from previous months. Do NOT submit zero amounts.
- Tax provision: 22% of (total revenue minus total expenses including depreciation). Debit 8700, credit 2920. Get balances AFTER all other closing entries to calculate correctly.

### Expense Analysis (identify accounts with largest increase)
When asked to find expense accounts with the largest INCREASE between two periods:
1. Call get_account_balances TWICE — once for each period separately:
   - get_account_balances(dateFrom="2026-01-01", dateTo="2026-01-31") for January
   - get_account_balances(dateFrom="2026-02-01", dateTo="2026-02-28") for February
2. Calculate the INCREASE per account: Feb balance - Jan balance
3. Sort by increase and pick ONLY the top N accounts (usually 3). Do NOT create projects for ALL accounts.
4. Create internal projects with isInternal=true for each of the top N
5. Create an activity for EACH project (create_project_activity)

### Overdue Invoice + Reminder Fee (purregebyr/Mahngebühr/taxa de lembrete)
1. find_entity(entityType="customer") to find the customer
2. find_invoice(customerId=X) to get ALL invoices for the customer
3. Identify the OVERDUE invoice: the one with invoiceDueDate BEFORE today's date. Do NOT pick the first invoice — check the due dates!
4. create_voucher: debit 1500 (with customerId), credit 3400 — for the reminder fee amount ONLY
5. create_invoice: create a NEW invoice for the reminder fee. Purregebyr/reminder fees are VAT-EXEMPT (vatTypeId=5, 0% MVA).
6. send_invoice: SEND the reminder invoice
7. register_payment: register partial payment on the OVERDUE invoice ID (from step 3), NOT the reminder invoice

### Error Correction in Ledger
1. get_account_balances (to see which accounts have wrong balances — filter by date range)
2. find_postings for the affected account (to find the wrong posting and its voucher ID)
3. For EACH error, find the COUNTERPART by looking up other postings on the same voucher:
   - Use find_postings or generic_get("ledger/posting", {"voucherId": X}) to find ALL postings on the voucher
   - The counterpart is the OTHER account on the same voucher (not the one with the error)
4. Create correcting vouchers using the ORIGINAL counterpart:
   - Wrong account: credit wrong account, debit correct account (same counterpart — usually 2400 for supplier)
   - Duplicate: credit the expense account, debit the ORIGINAL counterpart from the voucher (e.g. 2400)
   - Missing VAT: debit 2710 (input VAT), credit the expense account to reduce to net
   - Wrong amount: credit the difference on the expense account, debit the difference on the original counterpart
CRITICAL: Do NOT use 1920 (bank) or 9990 as counterpart. Look up the actual counterpart from the original voucher. It is usually 2400 (leverandørgjeld) for expense vouchers.

### Register Supplier/Vendor Invoice (leverandørfaktura)
1. find_entity (entityType="supplier") to find the supplier, or create_supplier if new
2. create_supplier_invoice (supplierId, invoiceNumber, invoiceDate, dueDate, amountInclVat, accountNumber)
   — Do NOT use create_voucher for supplier invoices. Use create_supplier_invoice.

### Year-End Closing (årsoppgjør/årsavslutning)
For "simplified year-end" (forenklet årsoppgjør) with SPECIFIC depreciation instructions:
- Do NOT use year_end_closing tool — it closes ALL accounts which is wrong for simplified closing
- Create SEPARATE vouchers for each depreciation, prepaid reversal, and tax provision as described in the prompt
- Use create_voucher for each entry individually

For FULL year-end closing of all accounts: use year_end_closing(date) tool.

## FILE HANDLING
- PDF invoices: text is auto-extracted. Look for: Fakturanummer (invoice number), Forfallsdato (due date), Beløp inkl. mva (amount incl. VAT), MVA (VAT), Leverandør (supplier).
- CSV bank statements: Try import_bank_statement first. If it fails, process rows manually (see Bank Reconciliation workflow above).
- Images: auto-included for visual analysis. Extract all text, numbers, dates, and amounts.

## RECEIPTS (kvittering)
When the prompt says to register an expense from a receipt (NOT "reiseregning"):
- The receipt lists items with prices. The price per item is EXCLUDING VAT. Total includes VAT.
- The prompt specifies WHICH item to register. Register ONLY that item.
- Choose the RIGHT method based on what the item is:
  a) Travel-related (flybillett, taxi, hotell, tog): use create_travel_expense with cost category
  b) Non-travel purchases (whiteboard, office supplies, dinner/representasjon, equipment): use create_voucher
     - Debit the correct expense account (e.g. 7300 for representasjon, 6500 for office supplies, 6560 for IT equipment)
     - Debit 2710 for input VAT (25% of the item price)
     - Credit 1920 (bank/bedriftskort) for the total including VAT
- If the prompt specifies a department (e.g. "HR", "Markedsføring", "IT"):
  - Find department ID with find_entity(entityType="department")
  - For travel_expense: pass departmentId parameter
  - For voucher: add "department": departmentId on the EXPENSE posting (not VAT or bank posting)
- Amounts: item price from receipt (excl. VAT), VAT = price × 0.25, total = price × 1.25
"""

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "find_employee",
            "description": "Find an existing employee by email or name. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "email": {"type": "string"},
                    "firstName": {"type": "string"},
                    "lastName": {"type": "string"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_employee",
            "description": "Create employee with employment record AND employment details (salary, STYRK, percentage, etc). Handles department lookup/creation automatically. Pass ALL details from the prompt/contract. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "firstName": {"type": "string"},
                    "lastName": {"type": "string"},
                    "email": {"type": "string"},
                    "dateOfBirth": {"type": "string", "description": "YYYY-MM-DD"},
                    "startDate": {"type": "string", "description": "Employment start date YYYY-MM-DD (tiltredelse)"},
                    "userType": {"type": "string", "enum": ["STANDARD", "EXTENDED", "NO_ACCESS"], "description": "EXTENDED for admin/kontoadministrator"},
                    "phoneNumberMobile": {"type": "string"},
                    "phoneNumberWork": {"type": "string"},
                    "phoneNumberHome": {"type": "string"},
                    "addressLine1": {"type": "string", "description": "Street address"},
                    "postalCode": {"type": "string"},
                    "city": {"type": "string"},
                    "country": {"type": "string", "description": "Country name"},
                    "nationalIdentityNumber": {"type": "string", "description": "Norwegian fødselsnummer/personnummer (11 digits)"},
                    "bankAccountNumber": {"type": "string", "description": "Norwegian bank account number (11 digits)"},
                    "departmentName": {"type": "string", "description": "Department name (e.g. 'Innkjøp', 'Salg'). Auto-creates if not exists."},
                    "annualSalary": {"type": "number", "description": "Annual salary (årslønn) in NOK"},
                    "monthlySalary": {"type": "number", "description": "Monthly salary (månedslønn) — auto-converts to annual"},
                    "hourlySalary": {"type": "number", "description": "Hourly wage (timelønn)"},
                    "percentageOfFullTimeEquivalent": {"type": "number", "description": "Employment percentage (stillingsprosent), e.g. 100.0, 80.0, 50.0"},
                    "occupationCode": {"type": "string", "description": "STYRK occupation code, e.g. '1211', '2411'"},
                    "employmentForm": {"type": "string", "description": "Fast stilling=PERMANENT, Midlertidig=TEMPORARY, Vikariat=TEMPORARY"},
                    "remunerationType": {"type": "string", "description": "Fastlønn=MONTHLY_WAGE, Timelønn=HOURLY_WAGE, Provisjon=COMMISION_PERCENTAGE"},
                    "employmentType": {"type": "string", "description": "ORDINARY (default), MARITIME, FREELANCE"},
                },
                "required": ["firstName", "lastName", "email"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_customer",
            "description": "Find an existing customer by name or organization number. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "organizationNumber": {"type": "string"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_customer",
            "description": "Create a new customer with optional address. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "email": {"type": "string"},
                    "invoiceEmail": {"type": "string", "description": "Invoice email address (defaults to email if not set)"},
                    "organizationNumber": {"type": "string"},
                    "phoneNumber": {"type": "string"},
                    "addressLine1": {"type": "string", "description": "Street address"},
                    "postalCode": {"type": "string"},
                    "city": {"type": "string"},
                    "country": {"type": "string", "description": "Country name, e.g. 'Norge', 'Germany'"},
                    "discountPercentage": {"type": "number", "description": "Discount percentage for the customer"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_product",
            "description": "Create a product. IMPORTANT: Always set 'number' if a product number/code is given. Set 'vatTypeId' if a specific VAT rate is mentioned. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "priceExcludingVat": {"type": "number"},
                    "number": {"type": "string", "description": "Product number/code (e.g. '9026')"},
                    "description": {"type": "string"},
                    "vatTypeId": {"type": "integer", "description": "VAT type: 3=25% (default), 31=15% (food/næringsmiddel), 32=12%, 5=0%"},
                    "costPrice": {"type": "number", "description": "Cost price excluding VAT"},
                    "productUnit": {"type": "string", "description": "Unit name or abbreviation (e.g. 'stk', 'kg', 'timer'). Looked up by name."},
                    "ean": {"type": "string", "description": "EAN/barcode number"},
                    "supplierId": {"type": "integer", "description": "Supplier ID to link the product to"},
                },
                "required": ["name", "priceExcludingVat"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_invoice",
            "description": "Find invoices for a customer. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "customerId": {"type": "integer"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_invoice",
            "description": "Create an invoice for a customer. Auto-resolves customer by name/org if customerId not provided. Handles bank account setup, product lookup, and order creation automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "customerId": {"type": "integer", "description": "Customer ID. Or provide customerName/organizationNumber instead."},
                    "customerName": {"type": "string", "description": "Customer name — auto-looks up customerId"},
                    "organizationNumber": {"type": "string", "description": "Org number — auto-looks up customerId"},
                    "invoiceDate": {"type": "string", "description": "YYYY-MM-DD (default: today)"},
                    "invoiceDueDate": {"type": "string", "description": "YYYY-MM-DD (default: invoiceDate + 14 days)"},
                    "kid": {"type": "string", "description": "KID number (kundeidentifikasjon) for the invoice"},
                    "invoiceRemark": {"type": "string", "description": "Invoice remarks (printed on invoice)"},
                    "invoiceComment": {"type": "string", "description": "Internal comment on the invoice"},
                    "projectId": {"type": "integer", "description": "Project ID to link the invoice to (for project billing)"},
                    "lines": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "description": {"type": "string", "description": "Line item description"},
                                "count": {"type": "number", "description": "Quantity (default 1). This is the NUMBER OF ITEMS, not a product code."},
                                "unitPrice": {"type": "number", "description": "Price per unit excluding VAT in NOK"},
                                "productNumber": {"type": "string", "description": "Product number/code to look up (e.g. '6390'). NOT the count."},
                                "productId": {"type": "integer", "description": "Tripletex product ID if already known"},
                                "vatTypeId": {"type": "integer", "description": "VAT type: 3=25%, 31=15%, 32=12%, 5=0%"},
                                "vatPercent": {"type": "number", "description": "VAT percentage (25, 15, 12, 0) — auto-converted to vatTypeId"},
                                "discount": {"type": "number", "description": "Discount percentage (0-100). For milestone billing: set unitPrice=full price, discount=remaining%. E.g. 50% milestone of 200000 → unitPrice=200000, discount=50."},
                                "currency": {"type": "string", "description": "Currency code (e.g. 'EUR', 'USD', 'SEK'). Default NOK."},
                            }
                        },
                        "description": "Invoice line items. IMPORTANT: 'count' is quantity, NOT product code. Numbers in parentheses like (6390) are product numbers, use 'productNumber' for those. For fixed price milestones: unitPrice=FULL price, discount=remaining%."
                    },
                },
                "required": ["lines"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_credit_note",
            "description": "Create a credit note to reverse an invoice. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer", "description": "Invoice ID to reverse (use find_invoice first)"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_project",
            "description": "Find an existing project by name. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_project",
            "description": "Create a new project. projectManagerId defaults to logged-in user. startDate defaults to today. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "projectManagerId": {"type": "integer", "description": "Employee ID of project manager (default: logged-in user)"},
                    "startDate": {"type": "string", "description": "YYYY-MM-DD (default: today)"},
                    "endDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "customerId": {"type": "integer"},
                    "description": {"type": "string", "description": "Project description"},
                    "fixedprice": {"type": "number", "description": "Fixed price amount. Also sets isFixedPrice=true"},
                    "isInternal": {"type": "boolean", "description": "Whether this is an internal project"},
                    "isFixedPrice": {"type": "boolean", "description": "Whether this is a fixed-price project"},
                    "projectCategoryId": {"type": "integer", "description": "Project category ID"},
                    "invoiceComment": {"type": "string", "description": "Default invoice comment for the project"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_travel_expense",
            "description": "Create a travel expense with costs, per diems, and/or mileage. Handles categories and payment types automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer", "description": "Employee ID (use find_employee first)"},
                    "title": {"type": "string", "description": "Title of the travel expense"},
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                    "destination": {"type": "string", "description": "Travel destination city (e.g. 'Trondheim')"},
                    "departureFrom": {"type": "string", "description": "Departure city (default: Oslo)"},
                    "projectId": {"type": "integer", "description": "Link to a project"},
                    "isForeignTravel": {"type": "boolean", "description": "Whether this is a foreign/international travel"},
                    "costs": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "description": {"type": "string"},
                                "amount": {"type": "number", "description": "Amount in NOK INCLUDING VAT"},
                                "categoryKeyword": {"type": "string", "description": "fly, taxi, diett, hotell, tog, parkering, buss, bil, annet"},
                            }
                        },
                        "description": "Receipted expenses (flights, taxi, etc)"
                    },
                    "perDiems": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "count": {"type": "integer", "description": "Number of days"},
                                "rate": {"type": "number", "description": "Daily rate in NOK"},
                                "location": {"type": "string"},
                                "overnightAccommodation": {"type": "string", "enum": ["NONE", "HOTEL", "BOARDING_HOUSE_WITHOUT_COOKING", "BOARDING_HOUSE_WITH_COOKING"]},
                                "isDeductionForBreakfast": {"type": "boolean"},
                                "isDeductionForLunch": {"type": "boolean"},
                                "isDeductionForDinner": {"type": "boolean"},
                            }
                        },
                        "description": "Per diem / daily allowance (dagdiett/døgndiett/Tagegeld)"
                    },
                    "mileageAllowances": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "km": {"type": "number", "description": "Distance in km"},
                                "rate": {"type": "number", "description": "Rate per km in NOK"},
                                "date": {"type": "string", "description": "YYYY-MM-DD"},
                                "departureLocation": {"type": "string"},
                                "destination": {"type": "string"},
                                "isCompanyCar": {"type": "boolean"},
                            }
                        },
                        "description": "Mileage/driving allowance (kjøregodtgjørelse)"
                    },
                    "accommodationAllowances": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "count": {"type": "integer", "description": "Number of nights"},
                                "rate": {"type": "number", "description": "Rate per night in NOK"},
                                "location": {"type": "string", "description": "Accommodation location"},
                            }
                        },
                        "description": "Accommodation allowance (nattillegg/losji)"
                    },
                },
                "required": ["employeeId", "title", "date"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_travel_expense",
            "description": "Delete a travel expense. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer"},
                    "title": {"type": "string", "description": "Title to match (partial match)"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "register_timesheet",
            "description": "Register timesheet hours for an employee on a project activity. Auto-adds employee as participant and creates activity if none exists. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer", "description": "Employee ID (use find_employee first)"},
                    "projectId": {"type": "integer", "description": "Project ID (use find_project first)"},
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                    "hours": {"type": "number"},
                    "activityName": {"type": "string", "description": "Name of the activity to match"},
                    "comment": {"type": "string"},
                    "hourlyRate": {"type": "number", "description": "Hourly rate override for this entry"},
                    "chargeable": {"type": "boolean", "description": "Whether this entry is chargeable"},
                },
                "required": ["employeeId", "projectId", "date", "hours"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_department",
            "description": "Create a department. Use sequential numbers for departmentNumber (1, 2, 3...), NOT the department name. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "departmentNumber": {"type": "string", "description": "Numeric sequence: '1', '2', '3'. NOT the department name."},
                    "departmentManagerId": {"type": "integer", "description": "Employee ID of the department manager"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_supplier",
            "description": "Create a supplier with optional address. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "email": {"type": "string"},
                    "invoiceEmail": {"type": "string", "description": "Invoice email address (defaults to email if not set)"},
                    "organizationNumber": {"type": "string"},
                    "phoneNumber": {"type": "string"},
                    "addressLine1": {"type": "string"},
                    "postalCode": {"type": "string"},
                    "city": {"type": "string"},
                    "country": {"type": "string"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_contact",
            "description": "Create a contact on a customer. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "firstName": {"type": "string"},
                    "lastName": {"type": "string"},
                    "customerId": {"type": "integer", "description": "Customer ID (use find_customer first)"},
                    "email": {"type": "string"},
                    "phoneNumberMobile": {"type": "string"},
                    "phoneNumberWork": {"type": "string", "description": "Work phone number"},
                },
                "required": ["firstName", "lastName", "customerId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_employee",
            "description": "Update an existing employee's contact info or details. Gets current version automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer"},
                    "firstName": {"type": "string"},
                    "lastName": {"type": "string"},
                    "email": {"type": "string"},
                    "dateOfBirth": {"type": "string", "description": "YYYY-MM-DD"},
                    "phoneNumberMobile": {"type": "string"},
                    "phoneNumberWork": {"type": "string"},
                    "phoneNumberHome": {"type": "string"},
                },
                "required": ["employeeId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_customer",
            "description": "Update an existing customer's info. Gets current version automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "customerId": {"type": "integer"},
                    "name": {"type": "string"},
                    "email": {"type": "string"},
                    "phoneNumber": {"type": "string"},
                    "organizationNumber": {"type": "string"},
                    "addressLine1": {"type": "string"},
                    "postalCode": {"type": "string"},
                    "city": {"type": "string"},
                    "country": {"type": "string"},
                },
                "required": ["customerId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "book_exchange_difference",
            "description": "Register payment on a foreign currency invoice AND book the exchange difference (agio/disagio). Handles the full flow: gets invoice, registers payment, calculates diff, creates voucher. Use this instead of manually calculating exchange differences.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer", "description": "Invoice ID (use find_invoice first)"},
                    "paymentAmountNOK": {"type": "number", "description": "Actual NOK received: foreign amount × CURRENT rate. E.g. 4868 EUR × 10.84 = 52769.12"},
                    "originalAmountNOK": {"type": "number", "description": "Original NOK when invoice was sent: foreign amount × ORIGINAL rate. E.g. 4868 EUR × 11.51 = 56030.68"},
                    "paymentDate": {"type": "string", "description": "YYYY-MM-DD (default: today)"},
                    "exchangeAccount": {"type": "integer", "description": "Exchange diff account (default: 7790)"},
                },
                "required": ["invoiceId", "paymentAmountNOK", "originalAmountNOK"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "register_payment",
            "description": "Register a payment on an invoice. For FULL payment, omit paymentAmount — the wrapper reads the exact amount from the invoice. Only set paymentAmount for PARTIAL payments. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer", "description": "Invoice ID (use find_invoice first)"},
                    "paymentDate": {"type": "string", "description": "YYYY-MM-DD (default: today)"},
                    "paymentAmount": {"type": "number", "description": "ONLY for partial payments. Omit for full payment — amount is read from the invoice automatically."},
                    "paymentTypeId": {"type": "integer", "description": "Override payment type ID (default: auto-detected)"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "reverse_payment",
            "description": "Reverse/undo a payment on an invoice (e.g. bank return). Finds the payment voucher and reverses it so the invoice shows outstanding amount again. Do NOT use create_credit_note for this — credit notes cancel the invoice, reverse_payment only undoes the payment. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer", "description": "Invoice ID (use find_invoice first)"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "grant_admin_role",
            "description": "Grant administrator entitlements to an employee. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer"},
                },
                "required": ["employeeId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "enable_department_accounting",
            "description": "Enable the department accounting module for the company. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {},
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "delete_entity",
            "description": "Delete any entity by type and ID. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "entityType": {"type": "string", "enum": ["employee", "customer", "product", "invoice", "order", "travelExpense", "project", "department", "supplier", "contact", "voucher"]},
                    "entityId": {"type": "integer"},
                },
                "required": ["entityType", "entityId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_travel_expense",
            "description": "Update an existing travel expense's title or date. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "travelExpenseId": {"type": "integer"},
                    "title": {"type": "string"},
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                },
                "required": ["travelExpenseId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_project",
            "description": "Update an existing project (e.g. set fixed price, change manager, change dates). Gets current version automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "projectId": {"type": "integer"},
                    "name": {"type": "string"},
                    "projectManagerId": {"type": "integer", "description": "Employee ID of project manager"},
                    "fixedprice": {"type": "number", "description": "Fixed price amount. Also sets isFixedPrice=true"},
                    "isFixedPrice": {"type": "boolean"},
                    "startDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "endDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "description": {"type": "string"},
                },
                "required": ["projectId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_accounts",
            "description": "Query the chart of accounts (ledger/account). Search by account number or name. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "number": {"type": "integer", "description": "Account number, e.g. 1920, 3000, 5000"},
                    "name": {"type": "string"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_postings",
            "description": "Query ledger postings. Filter by date range, account number, supplier, or customer. Use accountNumber (e.g. 6300), NOT accountId. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "dateFrom": {"type": "string", "description": "YYYY-MM-DD (default 2020-01-01)"},
                    "dateTo": {"type": "string", "description": "YYYY-MM-DD (default 2030-12-31)"},
                    "accountNumber": {"type": "integer", "description": "Account number (e.g. 1920, 6300, 7100). Preferred over accountId."},
                    "accountId": {"type": "integer", "description": "Tripletex internal ID. Use accountNumber instead if you have the number."},
                    "supplierId": {"type": "integer"},
                    "customerId": {"type": "integer"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_voucher",
            "description": "Create a ledger voucher with postings. Uses account NUMBERS (e.g. 1920, 3000) — automatically looks up IDs. IMPORTANT: Account 1500 (Kundefordringer) REQUIRES customerId. Account 2400 (Leverandørgjeld) REQUIRES supplierId. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                    "description": {"type": "string"},
                    "postings": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "accountNumber": {"type": "integer", "description": "Account number, e.g. 1920, 3000"},
                                "debitAmount": {"type": "number", "description": "Debit amount (positive)"},
                                "creditAmount": {"type": "number", "description": "Credit amount (positive)"},
                                "description": {"type": "string"},
                                "customerId": {"type": "integer", "description": "REQUIRED for account 1500 (Kundefordringer). Use find_customer to get the ID."},
                                "supplierId": {"type": "integer", "description": "REQUIRED for account 2400 (Leverandørgjeld). Use find_entity(supplier) to get the ID."},
                                "customDimension1": {"type": "string", "description": "Free accounting dimension 1 value name (e.g. 'Vestlandet')"},
                                "customDimension2": {"type": "string", "description": "Free accounting dimension 2 value name"},
                                "customDimension3": {"type": "string", "description": "Free accounting dimension 3 value name"},
                            }
                        }
                    },
                },
                "required": ["date", "description", "postings"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_entity",
            "description": "Find any entity by type and search params. Fallback when specific find_ actions don't exist. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "entityType": {"type": "string", "enum": ["employee", "customer", "product", "invoice", "order", "travelExpense", "project", "department", "supplier", "contact", "voucher"]},
                    "searchParams": {"type": "object", "description": "Search parameters like {name:'X', email:'Y'}"},
                },
                "required": ["entityType"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "run_payroll",
            "description": "Run payroll / create salary transaction for an employee. Handles salary types automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "employeeId": {"type": "integer", "description": "Employee ID (use find_employee first)"},
                    "date": {"type": "string", "description": "Payroll date YYYY-MM-DD"},
                    "year": {"type": "integer", "description": "Payroll year"},
                    "month": {"type": "integer", "description": "Payroll month (1-12)"},
                    "salaryLines": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "description": {"type": "string", "description": "e.g. 'Månedslønn', 'Bonus', 'Overtid'"},
                                "amount": {"type": "number", "description": "Amount in NOK"},
                                "count": {"type": "number", "description": "Count/quantity if applicable"},
                                "rate": {"type": "number", "description": "Rate if applicable"},
                            }
                        },
                        "description": "Salary line items (base salary, bonuses, etc)"
                    },
                },
                "required": ["employeeId", "date", "salaryLines"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_supplier_invoice",
            "description": "Register an incoming supplier/vendor invoice. Use this for 'leverandørfaktura', 'incoming invoice', 'factura del proveedor'. NOT ledger/voucher. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "supplierId": {"type": "integer", "description": "Supplier ID (use find_entity with entityType='supplier' first)"},
                    "invoiceNumber": {"type": "string", "description": "Invoice number from the supplier"},
                    "invoiceDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "dueDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "amountInclVat": {"type": "number", "description": "Total amount INCLUDING VAT"},
                    "accountNumber": {"type": "integer", "description": "Expense account number, e.g. 6590, 6300, 4000"},
                    "description": {"type": "string"},
                    "vatTypeId": {"type": "integer", "description": "1=25% input VAT (default), 11=15%, 12=12%, 0=no VAT"},
                    "currency": {"type": "string", "description": "Currency code (e.g. 'EUR', 'USD'). Default NOK."},
                    "kid": {"type": "string", "description": "KID number for the supplier invoice"},
                    "projectId": {"type": "integer", "description": "Link supplier invoice to a project"},
                },
                "required": ["supplierId", "invoiceNumber", "invoiceDate", "dueDate", "amountInclVat", "accountNumber"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "import_bank_statement",
            "description": "Import a bank statement file (CSV/CAMT/MT940) for bank reconciliation. Handles bankId and accountId lookup automatically. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "fileContent": {"type": "string", "description": "Raw file content as text (from attached CSV file)"},
                    "filename": {"type": "string", "description": "Filename, e.g. 'bank_statement.csv'"},
                    "fromDate": {"type": "string", "description": "Statement start date YYYY-MM-DD"},
                    "toDate": {"type": "string", "description": "Statement end date YYYY-MM-DD"},
                    "fileFormat": {"type": "string", "description": "DNB_CSV, NORDEA_CSV, DANSKE_BANK_CSV, etc. Default: DNB_CSV"},
                },
                "required": ["fileContent"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_bank_reconciliation",
            "description": "Create a bank reconciliation for a ledger account (e.g. account 1920). Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "accountId": {"type": "integer", "description": "Ledger account ID (use find_accounts first)"},
                },
                "required": ["accountId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "find_bank_transactions",
            "description": "Find bank statement transactions. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "bankStatementId": {"type": "integer"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "invoice_order",
            "description": "Convert an existing order into an invoice. Use when an order already exists and needs to be invoiced. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "orderId": {"type": "integer"},
                    "invoiceDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "sendToCustomer": {"type": "boolean", "description": "Send invoice to customer after creation"},
                },
                "required": ["orderId", "invoiceDate"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_project_activity",
            "description": "Create an activity, optionally linked to a project. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "name": {"type": "string"},
                    "projectId": {"type": "integer", "description": "Link activity to this project"},
                    "isChargeable": {"type": "boolean", "description": "Default true"},
                },
                "required": ["name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_travel_expense_vouchers",
            "description": "Book a travel expense in the ledger (create vouchers from it). Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "travelExpenseId": {"type": "integer"},
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                },
                "required": ["travelExpenseId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "pay_supplier_invoice",
            "description": "Register payment on a supplier/vendor invoice. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer", "description": "Supplier invoice ID"},
                    "paymentDate": {"type": "string", "description": "YYYY-MM-DD"},
                    "amount": {"type": "number", "description": "Payment amount (omit for full)"},
                    "paymentType": {"type": "integer", "description": "0 = auto-select (default)"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "approve_supplier_invoice",
            "description": "Approve a supplier invoice. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer"},
                    "comment": {"type": "string"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_opening_balance",
            "description": "Create an opening balance with account postings. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "YYYY-MM-DD"},
                    "postings": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "accountNumber": {"type": "integer"},
                                "debitAmount": {"type": "number"},
                                "creditAmount": {"type": "number"},
                            }
                        }
                    },
                },
                "required": ["date", "postings"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_ledger_account",
            "description": "Create a new ledger account in the chart of accounts. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "number": {"type": "integer", "description": "Account number"},
                    "name": {"type": "string", "description": "Account name"},
                },
                "required": ["number", "name"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "create_accounting_dimension",
            "description": "Create a free accounting dimension (fri regnskapsdimensjon) with values. Use for 'Prosjekttype', 'Kostsenter', etc. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "dimensionName": {"type": "string", "description": "Name of the dimension, e.g. 'Prosjekttype'"},
                    "description": {"type": "string"},
                    "values": {
                        "type": "array",
                        "items": {"type": "string"},
                        "description": "List of dimension values, e.g. ['Utvikling', 'Internt']"
                    },
                },
                "required": ["dimensionName"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "send_invoice",
            "description": "Send an invoice to the customer via email or other method. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "invoiceId": {"type": "integer"},
                    "sendType": {"type": "string", "enum": ["EMAIL", "EHF", "EFAKTURA", "LETTER", "MANUAL"], "description": "Default: EMAIL"},
                    "overrideEmailAddress": {"type": "string"},
                },
                "required": ["invoiceId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "approve_travel_expense",
            "description": "Approve a travel expense. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "travelExpenseId": {"type": "integer"},
                },
                "required": ["travelExpenseId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "deliver_travel_expense",
            "description": "Deliver/submit a travel expense for approval. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "travelExpenseId": {"type": "integer"},
                },
                "required": ["travelExpenseId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "reverse_voucher",
            "description": "Reverse a voucher (creates a reversing entry). Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "voucherId": {"type": "integer"},
                    "date": {"type": "string", "description": "YYYY-MM-DD (defaults to today)"},
                },
                "required": ["voucherId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "add_project_participant",
            "description": "Add an employee as participant to a project. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "projectId": {"type": "integer"},
                    "employeeId": {"type": "integer"},
                },
                "required": ["projectId", "employeeId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "suggest_bank_matches",
            "description": "Suggest transaction matches for a bank reconciliation. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "reconciliationId": {"type": "integer"},
                },
                "required": ["reconciliationId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "match_bank_transactions",
            "description": "Accept suggested matches and match remaining bank transactions for a reconciliation. Call AFTER suggest_bank_matches. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "reconciliationId": {"type": "integer"},
                },
                "required": ["reconciliationId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "close_bank_reconciliation",
            "description": "Close/finalize a bank reconciliation. Call AFTER matching transactions. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "reconciliationId": {"type": "integer"},
                },
                "required": ["reconciliationId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_account_balances",
            "description": "Get aggregated account balances from ledger postings. Returns per-account debit/credit totals. Useful for error correction and year-end closing. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "dateFrom": {"type": "string", "description": "YYYY-MM-DD (default 2020-01-01)"},
                    "dateTo": {"type": "string", "description": "YYYY-MM-DD (default 2030-12-31)"},
                    "accountNumberFrom": {"type": "integer", "description": "Filter: min account number (e.g. 3000)"},
                    "accountNumberTo": {"type": "integer", "description": "Filter: max account number (e.g. 7999)"},
                },
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "year_end_closing",
            "description": "Perform year-end closing (årsoppgjør). Automatically zeros out revenue (3xxx) and expense (4xxx-7xxx) accounts and transfers net to equity. Use this instead of manual voucher creation. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "date": {"type": "string", "description": "Closing date YYYY-MM-DD (e.g. 2025-12-31)"},
                    "fiscalYearStart": {"type": "string", "description": "Start of fiscal year YYYY-MM-DD (default Jan 1 of closing year)"},
                    "equityAccountNumber": {"type": "integer", "description": "Equity account for net transfer (default 8800)"},
                },
                "required": ["date"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_product",
            "description": "Update an existing product's name, price, or description. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "productId": {"type": "integer"},
                    "name": {"type": "string"},
                    "priceExcludingVat": {"type": "number"},
                    "description": {"type": "string"},
                },
                "required": ["productId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_supplier",
            "description": "Update an existing supplier's info. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "supplierId": {"type": "integer"},
                    "name": {"type": "string"},
                    "email": {"type": "string"},
                    "phoneNumber": {"type": "string"},
                    "organizationNumber": {"type": "string"},
                    "addressLine1": {"type": "string"},
                    "postalCode": {"type": "string"},
                    "city": {"type": "string"},
                    "country": {"type": "string"},
                },
                "required": ["supplierId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_contact",
            "description": "Update an existing contact's info. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "contactId": {"type": "integer"},
                    "firstName": {"type": "string"},
                    "lastName": {"type": "string"},
                    "email": {"type": "string"},
                    "phoneNumberMobile": {"type": "string"},
                    "phoneNumberWork": {"type": "string"},
                },
                "required": ["contactId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "update_department",
            "description": "Update an existing department's name or number. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "departmentId": {"type": "integer"},
                    "name": {"type": "string"},
                    "departmentNumber": {"type": "string"},
                },
                "required": ["departmentId"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_tool_fields",
            "description": "Get ALL available fields for a Tripletex entity type. NO API calls — reads local spec. Call BEFORE create/update to see every field you can set.",
            "parameters": {
                "type": "object",
                "properties": {
                    "entityType": {"type": "string", "description": "Entity type: Employee, Customer, Product, Invoice, Order, OrderLine, Project, Department, Supplier, Contact, TravelExpense, Voucher, Posting, SalaryTransaction, Address, Cost, PerDiemCompensation, MileageAllowance, Employment"},
                },
                "required": ["entityType"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "search_api",
            "description": "Search the Tripletex API spec for endpoints. NO API calls — reads local OpenAPI spec. Use to discover endpoints for unfamiliar operations. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "query": {"type": "string", "description": "Search term, e.g. 'payment', 'salary', 'reconciliation', 'module'"},
                },
                "required": ["query"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_api_endpoint",
            "description": "Get full documentation for a specific API endpoint path. NO API calls — reads local OpenAPI spec. Returns all parameters, request body fields, types. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "API path, e.g. '/employee', '/invoice', '/salary/transaction'"},
                    "method": {"type": "string", "description": "HTTP method: GET, POST, PUT, DELETE. Omit to see all methods."},
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generic_get",
            "description": "GET any Tripletex endpoint. Use for any read operation. Use search_api + get_api_endpoint first to find the right path and params. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "API path, e.g. 'employee', 'invoice', 'ledger/posting'"},
                    "params": {"type": "object", "description": "Query parameters as key-value pairs"},
                },
                "required": ["path"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generic_post",
            "description": "POST to any Tripletex endpoint. Use for any create operation not covered by specific tools. Use get_tool_fields or get_api_endpoint first to see required fields. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "API path, e.g. 'employee', 'customer', 'ledger/voucher'"},
                    "body": {"type": "object", "description": "JSON request body with all required fields"},
                },
                "required": ["path", "body"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generic_put",
            "description": "PUT to update any entity or call action endpoints. For :action endpoints (e.g. invoice/{id}/:payment), params are sent as query params. For regular PUT, body is sent as JSON. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "API path with ID, e.g. 'employee/123' or 'invoice/456/:payment'"},
                    "body": {"type": "object", "description": "JSON body (regular PUT) or query params (:action endpoints)"},
                },
                "required": ["path", "body"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "generic_delete",
            "description": "DELETE any entity by path+ID. Call get_tool_fields() for complete field list.",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "API path with ID, e.g. 'employee/123', 'travelExpense/456'"},
                },
                "required": ["path"]
            }
        }
    },
]


def extract_file_content(files: list[dict]) -> tuple[str, list[dict]]:
    """Extract text content from attached files with robust error handling.

    Returns:
        tuple of (extracted_text, image_parts)
        image_parts: list of {"type":"image_url","image_url":{"url":"data:mime;base64,..."}}
    """
    extracted = []
    image_parts = []

    for f in files:
        filename = f.get("filename", "unknown")
        mime = f.get("mime_type", "")
        raw_b64 = f.get("content_base64", "")

        # Safe base64 decode
        try:
            raw = base64.b64decode(raw_b64)
        except Exception as e:
            extracted.append(f"[File: {filename}] Error decoding base64: {e}")
            continue

        if not raw:
            extracted.append(f"[File: {filename}] Empty file (0 bytes)")
            continue

        if mime == "application/pdf" or filename.lower().endswith(".pdf"):
            try:
                pdf = pdfplumber.open(io.BytesIO(raw))
                parts = []
                for i, page in enumerate(pdf.pages):
                    text = page.extract_text() or ""
                    if text.strip():
                        parts.append(text)

                    tables = page.extract_tables()
                    for j, table in enumerate(tables):
                        if table:
                            table_text = "\n".join(
                                " | ".join(str(cell) if cell else "" for cell in row)
                                for row in table
                            )
                            parts.append(f"[Table {j+1} on page {i+1}]\n{table_text}")
                pdf.close()

                if parts:
                    extracted.append(f"[File: {filename} — extracted text]\n" + "\n\n".join(parts))
                else:
                    extracted.append(
                        f"[File: {filename}] PDF with no extractable text — likely scanned. "
                        f"See the raw PDF image below for visual content."
                    )

                # Also include raw PDF as image for visual cross-reference
                # Convert first page to image-like representation via base64
                # The LLM can't see PDFs directly but we note it
                extracted.append(
                    f"[File: {filename} — raw base64 available] "
                    f"If the extracted text above is incomplete, the original PDF ({len(raw)} bytes) "
                    f"may contain additional visual information not captured by text extraction."
                )
            except Exception as e:
                extracted.append(f"[File: {filename}] Error extracting PDF: {e}")

        elif mime.startswith("text/") or filename.lower().endswith((".csv", ".txt", ".json", ".xml")):
            text = None
            for encoding in ["utf-8", "latin-1", "iso-8859-1", "cp1252"]:
                try:
                    text = raw.decode(encoding)
                    break
                except (UnicodeDecodeError, ValueError):
                    continue
            if text is None:
                text = raw.decode("utf-8", errors="replace")
            extracted.append(f"[File: {filename}]\n{text}")

        elif filename.lower().endswith((".xlsx", ".xls")):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    rows = []
                    for row in ws.iter_rows(values_only=True):
                        rows.append(" | ".join(str(c) if c is not None else "" for c in row))
                    if rows:
                        extracted.append(f"[File: {filename} / Sheet: {sheet}]\n" + "\n".join(rows))
                wb.close()
            except ImportError:
                extracted.append(f"[File: {filename}] Excel file ({len(raw)} bytes) — openpyxl not installed")
            except Exception as e:
                extracted.append(f"[File: {filename}] Error reading Excel: {e}")

        elif mime.startswith("image/") or filename.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".webp", ".tiff", ".bmp")):
            # Include as vision content for the LLM
            img_mime = mime or "image/jpeg"
            image_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{img_mime};base64,{raw_b64}"},
            })
            extracted.append(
                f"[Image: {filename}] ({img_mime}, {len(raw)} bytes) — "
                f"Image included below for visual analysis. Extract all text, numbers, dates, and amounts from it."
            )

        else:
            extracted.append(f"[File: {filename}] Unsupported file type ({mime}, {len(raw)} bytes)")

    return "\n\n".join(extracted), image_parts


# Map tool names to action functions
ACTION_MAP = {
    "find_employee": actions.find_employee,
    "create_employee": actions.create_employee,
    "update_employee": actions.update_employee,
    "find_customer": actions.find_customer,
    "create_customer": actions.create_customer,
    "update_customer": actions.update_customer,
    "create_product": actions.create_product,
    "find_invoice": actions.find_invoice,
    "create_invoice": actions.create_invoice,
    "create_credit_note": actions.create_credit_note,
    "book_exchange_difference": actions.book_exchange_difference,
    "register_payment": actions.register_payment,
    "reverse_payment": actions.reverse_payment,
    "find_project": actions.find_project,
    "create_project": actions.create_project,
    "create_travel_expense": actions.create_travel_expense,
    "delete_travel_expense": actions.delete_travel_expense,
    "register_timesheet": actions.register_timesheet,
    "create_department": actions.create_department,
    "create_supplier": actions.create_supplier,
    "create_contact": actions.create_contact,
    "update_travel_expense": actions.update_travel_expense,
    "update_product": actions.update_product,
    "update_supplier": actions.update_supplier,
    "update_contact": actions.update_contact,
    "update_department": actions.update_department,
    "invoice_order": actions.invoice_order,
    "create_project_activity": actions.create_project_activity,
    "create_travel_expense_vouchers": actions.create_travel_expense_vouchers,
    "pay_supplier_invoice": actions.pay_supplier_invoice,
    "approve_supplier_invoice": actions.approve_supplier_invoice,
    "create_opening_balance": actions.create_opening_balance,
    "create_ledger_account": actions.create_ledger_account,
    "create_accounting_dimension": actions.create_accounting_dimension,
    "send_invoice": actions.send_invoice,
    "approve_travel_expense": actions.approve_travel_expense,
    "deliver_travel_expense": actions.deliver_travel_expense,
    "reverse_voucher": actions.reverse_voucher,
    "add_project_participant": actions.add_project_participant,
    "suggest_bank_matches": actions.suggest_bank_matches,
    "match_bank_transactions": actions.match_bank_transactions,
    "close_bank_reconciliation": actions.close_bank_reconciliation,
    "get_account_balances": actions.get_account_balances,
    "year_end_closing": actions.year_end_closing,
    "update_project": actions.update_project,
    "find_accounts": actions.find_accounts,
    "find_postings": actions.find_postings,
    "create_voucher": actions.create_voucher,
    "grant_admin_role": actions.grant_admin_role,
    "enable_department_accounting": actions.enable_department_accounting,
    "run_payroll": actions.run_payroll,
    "create_supplier_invoice": actions.create_supplier_invoice,
    "import_bank_statement": actions.import_bank_statement,
    "create_bank_reconciliation": actions.create_bank_reconciliation,
    "find_bank_transactions": actions.find_bank_transactions,
    "delete_entity": actions.delete_entity,
    "find_entity": actions.find_entity,
    "get_tool_fields": actions.get_tool_fields,
    "search_api": actions.search_api,
    "get_api_endpoint": actions.get_api_endpoint,
    "generic_get": actions.generic_get,
    "generic_post": actions.generic_post,
    "generic_put": actions.generic_put,
    "generic_delete": actions.generic_delete,
}


def execute_tool(tx: TripletexClient, name: str, args: dict, task_id: str = "") -> str:
    """Execute a tool call via deterministic action wrapper."""
    tag = f"[{task_id}] " if task_id else ""

    # Strip junk the LLM sends despite being told not to: empty strings, zeros for optional ID fields
    clean_args = {}
    for k, v in args.items():
        if isinstance(v, str) and v.strip() == "" and k not in ("prompt", "query", "description"):
            continue  # Skip empty strings (except for fields where empty is meaningful)
        if isinstance(v, (int, float)) and v == 0 and k.endswith("Id"):
            continue  # Skip zero IDs
        clean_args[k] = v
    args = clean_args

    try:
        action_fn = ACTION_MAP.get(name)
        if not action_fn:
            logger.error("%sUnknown action: %s", tag, name)
            return json.dumps({"error": f"Unknown action: {name}"})

        logger.info("%sTOOL CALL: %s\n  args: %s", tag, name,
                    json.dumps(args, ensure_ascii=False, default=str))

        result = action_fn(tx=tx, **args)

        # If the API returned an error, format it clearly for the LLM
        if isinstance(result, dict) and result.get("_is_error"):
            error_msg = result.get("_error_summary", "API call failed")
            result_json = json.dumps({
                "error": True,
                "summary": error_msg,
                "fix_hint": "Read the error details above. Fix the specific field mentioned and retry with corrected values in ONE attempt.",
                "raw": {k: v for k, v in result.items() if not k.startswith("_")},
            }, ensure_ascii=False, default=str)
            logger.warning("%sTOOL ERROR: %s\n  result: %s", tag, name, result_json)
            return result_json

        result_json = json.dumps(result, ensure_ascii=False, default=str)
        logger.info("%sTOOL OK: %s\n  result: %s", tag, name, result_json)
        return result_json
    except Exception as e:
        logger.error("%sTOOL CRASH: %s — %s", tag, name, str(e), exc_info=True)
        return json.dumps({
            "error": True,
            "summary": f"Action {name} crashed: {str(e)}",
            "fix_hint": "Check your arguments and try again with corrected values.",
        })


def solve_task(
    prompt: str,
    files: list[dict],
    base_url: str,
    session_token: str,
    openrouter_api_key: str | None = None,
    max_iterations: int = 25,
    task_id: str = "",
) -> dict:
    """Run the agentic loop to solve a Tripletex accounting task."""
    import datetime
    tag = f"[{task_id}] " if task_id else ""

    tx = TripletexClient(base_url, session_token)
    actions._reset_cache()
    api_key = openrouter_api_key or os.environ.get("OPENROUTER_API_KEY", "")

    client = OpenAI(
        base_url="https://openrouter.ai/api/v1",
        api_key=api_key,
    )

    # Build user message
    today = datetime.date.today().isoformat()
    text_content = f"## Task\nToday's date: {today}\n\n{prompt}"
    image_parts = []

    if files:
        file_text, image_parts = extract_file_content(files)
        text_content += f"\n\n## Attached Files (extracted text)\n{file_text}"
        logger.info("%sExtracted file text (%d chars):\n%s", tag, len(file_text), file_text)

    # Build user message — multipart if images, plain text otherwise
    if image_parts:
        user_message_content = [
            {"type": "text", "text": text_content},
            *image_parts,
        ]
        logger.info("%sUser message: text + %d image(s)", tag, len(image_parts))
    else:
        user_message_content = text_content

    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {"role": "user", "content": user_message_content},
    ]

    # Track all actions for the summary
    all_actions = []

    for iteration in range(max_iterations):
        logger.info("%s── Iteration %d/%d ──", tag, iteration + 1, max_iterations)

        response = client.chat.completions.create(
            model="openai/gpt-5.4-mini",
            messages=messages,
            tools=TOOLS,
            temperature=0.1,
        )

        choice = response.choices[0]
        msg = choice.message
        finish_reason = choice.finish_reason

        # Log LLM response text (if any)
        if msg.content:
            logger.info("%sLLM text (finish_reason=%s):\n%s", tag, finish_reason, msg.content)

        messages.append(msg)

        if not msg.tool_calls:
            logger.info("%sAgent finished after %d iterations (finish_reason=%s)", tag, iteration + 1, finish_reason)
            if finish_reason == "length":
                logger.warning("%sOutput was TRUNCATED — may have missed tool calls!", tag)

            # Check if agent only did lookups but never mutated anything — likely stopped early
            action_names = [a["name"] for a in all_actions]
            read_only_prefixes = ("find_", "generic_get", "get_account", "get_tool_fields", "search_api", "get_api_endpoint")
            did_mutate = any(
                not any(n.startswith(p) for p in read_only_prefixes)
                for n in action_names
            )
            if not did_mutate and iteration < max_iterations - 1:
                logger.warning("%sAgent stopped without mutating! Actions so far: %s — nudging to continue",
                             tag, action_names or "none")
                messages.append({
                    "role": "user",
                    "content": "You have not completed the task yet. Re-read the original task and perform the required create/update actions now."
                })
                continue

            # Log task summary
            _log_task_summary(tag, all_actions, iteration + 1, tx._call_count)
            return {"status": "completed", "iterations": iteration + 1}

        logger.info("%sLLM requested %d tool call(s) (finish_reason=%s)", tag, len(msg.tool_calls), finish_reason)

        for tc in msg.tool_calls:
            args = json.loads(tc.function.arguments)
            result = execute_tool(tx, tc.function.name, args, task_id=task_id)

            # Track for summary
            is_error = '"error": true' in result.lower() or '"error":true' in result.lower()
            all_actions.append({
                "name": tc.function.name,
                "args": args,
                "result_length": len(result),
                "is_error": is_error,
                "iteration": iteration + 1,
            })

            messages.append({
                "role": "tool",
                "tool_call_id": tc.id,
                "content": result,
            })

    logger.warning("%sHit max iterations (%d)!", tag, max_iterations)
    _log_task_summary(tag, all_actions, max_iterations, tx._call_count)
    return {"status": "completed", "iterations": max_iterations}


def _log_task_summary(tag: str, actions: list[dict], iterations: int, api_calls: int):
    """Log a structured summary of the entire task execution."""
    read_only_prefixes = ("find_", "generic_get", "get_account", "get_tool_fields", "search_api", "get_api_endpoint")
    errors = [a for a in actions if a["is_error"]]
    mutates = [a for a in actions
               if not any(a["name"].startswith(p) for p in read_only_prefixes)]

    lines = [
        f"{tag}{'=' * 60}",
        f"{tag}TASK SUMMARY",
        f"{tag}  Iterations: {iterations}",
        f"{tag}  Tool calls: {len(actions)}",
        f"{tag}  API calls:  {api_calls}",
        f"{tag}  Mutations:  {len(mutates)}",
        f"{tag}  Errors:     {len(errors)}",
        f"{tag}",
        f"{tag}  Action sequence:",
    ]
    for i, a in enumerate(actions):
        status = "ERROR" if a["is_error"] else "OK"
        args_summary = json.dumps(a["args"], ensure_ascii=False, default=str)
        if len(args_summary) > 200:
            args_summary = args_summary[:200] + "..."
        lines.append(f"{tag}    {i+1}. [{status}] {a['name']}({args_summary})")

    if errors:
        lines.append(f"{tag}")
        lines.append(f"{tag}  Failed actions:")
        for a in errors:
            lines.append(f"{tag}    - {a['name']} (iteration {a['iteration']})")

    lines.append(f"{tag}{'=' * 60}")
    logger.info("\n".join(lines))
